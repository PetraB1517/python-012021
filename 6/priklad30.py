"""Export do Excelu
pandas umí uložit data i do Excel sešitu, se kterým se bude uživatelům lépe pracovat.
K ukádání do Excelu využívá pandas modul openpyxl, který ale není nainstalován automaticky.
Nainstaluj si modul openpyxl standardním způsobem, který jsme si ukazovali v lekci.

Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor.
Pokud jsi tento příklad nezpracovala, ulož libovolnou jinou tabulku jako Excel sešit.
Výsledný sešit si otevři a zkontroluj. K uložení použij funkci to_excel, se kterou pracuj stejně,
jako jsme na lekci pracovali s funkci to_csv.

Zkus data z Excelu naopak načíst pomocí funkci read_excel() a ověř, že se soubor načetl v pořádku.
"""
import openpyxl
import pandas
vykazy = pandas.read_excel("vykazy.xlsx")
print(vykazy)

"""
Rozšířené zadání
Modul openpyxl lze používat i samostatně. Přehled jeho funkcí najdeš v dokumentaci.

Zkus pomocí modulu zapsat rozvrh hodin jako tabulku v Excel sešitu.
Níže máš program, který obsahuje rozvrh hodin jako dvourozměrný seznam.
Vnitřní seznamy obsahují předměty v rozvrhu, jeden vnitřní seznam vždy obsahuje předměty pro daný den.

Podívej se nejprve na ukázku.
Jednoduchý program níže zapíše hodnotu Test do buňky B1 a výsledek uloží souboru rozvrh_hodin.xlsx.
Pokud neznáš terminologii Excelu, pak Workbook označuje sešit, tj. celý "soubor". ws se odkazuje na Work Sheet,
což je list, tj. jedna "záložka". Náš soubor bude mít jen jeden list, čímž je situace jednoduchá.

from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

wb.save(filename="rozvrh_hodin.xlsx")

Projdi si nyní následující program a v programu doplň kód na zápis názvu předmětu do buňky.

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")


Dobrovolný doplněk rozšířeného zadání :-)
Modul openpyxl je silný v možnostech formátování, které nabízí. Podívej se na rozšířený příklad níže.
Ten kromě zápisu hodnoty nastaví buňce šedivou barvu jako pozadí.

Přehled barev, které můžeš použít, najdeš v dokumentaci.

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

# Určím si barvu
sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
# Přiřadím barvu buňce
bunka.fill = sediva_barva

wb.save(filename="rozvrh_hodin.xlsx")

Vrať se ke svému programu a zkus vytvořit rozvrh obarvený.
Například můžeš zkusit podbarvit oběd šedě a poté třeba nastavit nějakou barvu výchovám, jinou barvu jazykům atd.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]

sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
modra_barva = PatternFill("solid", fgColor="000066FF")
zluta_barva = PatternFill("solid", fgColor="00FFFF00")
cervena_barva = PatternFill("solid", fgColor="00CC3333")
zelena_barva = PatternFill("solid", fgColor="0000FF00")
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    if predmet == "Oběd":
      bunka.fill = sediva_barva
    elif "výchova" in predmet:
      bunka.fill = modra_barva
    elif "jazyk" in predmet:
      bunka.fill = zluta_barva
    elif ("Matematika" in predmet) or ("Fyzika" in predmet):
      bunka.fill = zelena_barva
    else:
      bunka.fill = cervena_barva
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")

